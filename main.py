from fastapi import FastAPI, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import json
import os
import uuid
from openpyxl import load_workbook
import tempfile
from datetime import datetime, timedelta
import requests
import openai

# --- TVŮJ MAPY.CZ a OpenAI KLÍČ ---
MAPYCZ_API_KEY = "bpImyVMf55mk8c8ayWQde_xazVqCbPb8YXAFdnfB8"
OPENAI_API_KEY = "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"
openai.api_key = OPENAI_API_KEY

CARS_JSON = [
  {
    "Firma": "OK BROKERS s.r.o.",
    "IČO": 25267001,
    "Vozidlo": "Škoda Superb",
    "RZ": "8M15169",
    "PHM": "Nafta",
    "Spotřeba l/100km": 6.23,
    "Řidič/odpovědná osoba": "Kubišová Jana"
  },
  {
    "Firma": "OK ECONOMY s.r.o.",
    "IČO": 44014929,
    "Vozidlo": "Škoda Superb",
    "RZ": "OKH00003",
    "PHM": "Benzín",
    "Spotřeba l/100km": 6.23,
    "Řidič/odpovědná osoba": "Šimková Petra"
  },
  {
    "Firma": "OK GROUP",
    "IČO": 28110056,
    "Vozidlo": "BMW 5",
    "RZ": "OKH00033",
    "PHM": "Benzín",
    "Spotřeba l/100km": 6.4,
    "Řidič/odpovědná osoba": "Ing. Michal Kubiš"
  },
  {
    "Firma": "Agroteam CZ s.r.o.",
    "IČO": 25561804,
    "Vozidlo": "Škoda Kodiaq",
    "RZ": "OKH30620",
    "PHM": "Benzín",
    "Spotřeba l/100km": 7.4,
    "Řidič/odpovědná osoba": "Kubišová Iva"
  },
  {
    "Firma": "OK Group a.s.",
    "IČO": 25561804,
    "Vozidlo": "BMW X5",
    "RZ": "7AA3030",
    "PHM": "Nafta",
    "Spotřeba l/100km": 9.9,
    "Řidič/odpovědná osoba": "Milan Ondra"
  },
  {
    "Firma": "OK GROUP s.r.o.",
    "IČO": 25561804,
    "Vozidlo": "ASTON MARTIN",
    "RZ": "OKH00007",
    "PHM": "Benzín",
    "Spotřeba l/100km": 10.2,
    "Řidič/odpovědná osoba": "Ing. Vladimíra Kubišová"
  },
  {
    "Firma": "OK Group a.s.,",
    "IČO": 25561804,
    "Vozidlo": "Škoda Superb",
    "RZ": "2BZ1888",
    "PHM": "Benzín",
    "Spotřeba l/100km": 6.23,
    "Řidič/odpovědná osoba": "Maloch Jan"
  },
  {
    "Firma": "OK Group a.s.",
    "IČO": 25561804,
    "Vozidlo": "Škoda Superb",
    "RZ": "2BZ1777",
    "PHM": "Benzín",
    "Spotřeba l/100km": 6.23,
    "Řidič/odpovědná osoba": "Malochová Renata"
  },
  {
    "Firma": "OK KLIENT a.s.",
    "IČO": 29185114,
    "Vozidlo": "BMW 5",
    "RZ": "OKH00088",
    "PHM": "Benzín",
    "Spotřeba l/100km": 6.4,
    "Řidič/odpovědná osoba": "Kubiš Radoslav ml."
  },
  {
    "Firma": "OK Group a.s.",
    "IČO": 25561804,
    "Vozidlo": "BMW 5",
    "RZ": "OKH00011",
    "PHM": "Natural",
    "Spotřeba l/100km": 6.4,
    "Řidič/odpovědná osoba": "Ing Kubiš Radoslav"
  },
  {
    "Firma": "Agroteam CZ s.r.o.",
    "IČO": 25561804,
    "Vozidlo": "LR Defender",
    "RZ": "OKH00001",
    "PHM": "Benzín",
    "Spotřeba l/100km": 11,
    "Řidič/odpovědná osoba": "Ing. Kubiš Radoslav"
  },
  {
    "Firma": "Agroteam CZ s.r.o.",
    "IČO": 25561804,
    "Vozidlo": "Range Rover",
    "RZ": "0KH11111",
    "PHM": "Benzín",
    "Spotřeba l/100km": 10.2,
    "Řidič/odpovědná osoba": "Ing. Kubiš Radoslav"
  },
  {
    "Firma": "OK GRANT s.r.o.",
    "IČO": 28268318,
    "Vozidlo": "BMW 8",
    "RZ": "0KH00002",
    "PHM": "Benzín",
    "Spotřeba l/100km": 7.4,
    "Řidič/odpovědná osoba": "Ing. Vladimíra Kubišová"
  }
]

EXCEL_TEMPLATE_PATH = "Template_Kniha_jizd_simulace.xlsx"

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/", response_class=HTMLResponse)
def read_root():
    with open("static/index.html", encoding="utf-8") as f:
        return f.read()

@app.get("/cars")
def get_cars():
    return CARS_JSON

def get_real_distance_km(from_addr, to_addr):
    url = (
        f"https://api.mapy.cz/route?"
        f"format=json&apikey={MAPYCZ_API_KEY}&from={from_addr}&to={to_addr}"
    )
    resp = requests.get(url)
    try:
        data = resp.json()
        return data["result"]["length"] / 1000  # metry na km
    except Exception:
        return 0

def prepare_ai_prompt(car, expanded_events, real_distances):
    return f"""
You are an expert assistant for generating vehicle trip logs (kniha jízd).
Based on the input below, generate a realistic list of trips for each day, in table format.
Each day must contain at least one trip, starting with the first date in events.

Strict rules:
- The trip route (column 'Trasa') is always the event location.
- The 'Počet km' column should match the real distance between previous and current event (see below).
- Refueling (tankování) is always written in column B if it occurs.
- Multi-day events (ubytování, jiné) must be listed for each day.
- 'Účel cesty' leave empty unless a note is present.
- Dates are always in dd.mm.yyyy format.
- The table columns are: Datum, Tankování, Trasa, Počet km, Účel cesty.
- If data is missing, leave the cell blank.

**Vehicle:**  
{json.dumps(car, ensure_ascii=False)}

**Events:**  
{json.dumps(expanded_events, ensure_ascii=False)}

**Real distances between points (km):**  
{json.dumps(real_distances, ensure_ascii=False)}

Output only the table as an array of rows in this format:
[
  ["01.04.2025", "Tankování: Brno OMV", "Brno OMV", 0, ""],
  ["02.04.2025", "", "Praha Vyšehrad", 210, ""],
  ...
]
""".strip()

@app.post("/simulate")
async def simulate(request: Request):
    data = await request.json()
    car = data["car"]
    events = data["events"]

    # 1. Rozbal vícedenní události na jednotlivé dny
    expanded_events = []
    for e in events:
        if e["typ"] in ["Ubytování", "Jiné"] and e.get("datum_konec"):
            start = datetime.strptime(e["datum"], "%Y-%m-%d")
            end = datetime.strptime(e["datum_konec"], "%Y-%m-%d")
            for i in range((end - start).days + 1):
                day = start + timedelta(days=i)
                expanded_events.append({
                    **e,
                    "datum": day.strftime("%Y-%m-%d"),
                    "datum_konec": None
                })
        else:
            expanded_events.append(e)

    # 2. Spočítej reálné vzdálenosti mezi po sobě jdoucími událostmi
    default_address = "Demlova 265/12, Černá Pole (Brno-sever), 613 00 Brno"
    real_distances = []
    prev_addr = default_address
    for e in expanded_events:
        curr_addr = e.get("misto", default_address)
        km = get_real_distance_km(prev_addr, curr_addr) if prev_addr != curr_addr else 0
        real_distances.append(km)
        prev_addr = curr_addr

    # 3. Připrav prompt pro OpenAI
    prompt = prepare_ai_prompt(car, expanded_events, real_distances)

    # 4. Zavolej OpenAI (GPT-4o)
    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=2000,
        temperature=0.05,
    )

    # 5. Zpracuj odpověď OpenAI (musí být pole pole, tedy [ [...], [...], ... ])
    import ast
    try:
        ai_table = ast.literal_eval(response.choices[0].message.content)
    except Exception:
        # fallback pokud je odpověď jiná (můžeš upravit debug výstup)
        ai_table = []

    # 6. Zapiš do excelu (od řádku 10)
    file_id = str(uuid.uuid4())
    temp_path = os.path.join(tempfile.gettempdir(), f"{file_id}.xlsx")
    wb = load_workbook(EXCEL_TEMPLATE_PATH)
    ws = wb.active

    ws["B3"] = car.get("Vozidlo", "")
    ws["B4"] = car.get("IČO", "")
    ws["B5"] = car.get("Vozidlo", "")
    ws["B6"] = car.get("RZ", "")
    ws["B7"] = car.get("PHM", "")
    ws["B8"] = car.get("Řidič/odpovědná osoba", "")

    row_start = 10
    for i, row in enumerate(ai_table):
        for j, cell in enumerate(row):
            col = chr(65 + j)  # A=65, B=66, C=67 ...
            ws[f"{col}{row_start + i}"] = cell

    # Dole 2 prázdné řádky, pak Datum a Podpis
    summary_row = row_start + len(ai_table) + 2
    ws[f"A{summary_row}"] = "Datum"
    ws[f"A{summary_row+1}"] = "Podpis"

    wb.save(temp_path)
    return {"success": True, "download_url": f"/download/{file_id}"}

@app.get("/download/{file_id}")
def download_excel(file_id: str):
    temp_path = os.path.join(tempfile.gettempdir(), f"{file_id}.xlsx")
    return FileResponse(temp_path, filename="Kniha_jizd.xlsx")
